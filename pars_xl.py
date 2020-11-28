from openpyxl import load_workbook

def get_words():
	wb = load_workbook("d_lang.xlsx")
	ws = wb.worksheets[3]

	words = {}

	rowLen = max((c.row for c in ws['A'] if c.value is not None))
	for row in ws.iter_rows(min_row=1, 		min_col=1, max_row=rowLen, max_col=2):
		eng = row[0].value
		dwa = row[1].value
		words[eng] = dwa
			
	wb.close()
	w = list(words)
	print(words)
	print(w)
	
get_words()